# -*- coding: utf-8 -*-
# Change log:
# 2026-09-01: Sannsynlighetsbegrunnelse from U/V pairs (not T question labels), same pattern as Q/R.
# 2026-09-01: Map Excel KIT paragraph (Konfidensialitet/Integritet/Tilgjengelighet) to K, I, T tags (varchar 50).
# 2026-09-01: Detect first numeric RiskID row; skip empty/header blocks (newer xlsm mal starts at row 6).
# 2026-09-01: Resolve Risikovurdering columns from headers (Y=tiltak, S/W levels) with letter fallback for old xlsx.
# 2026-09-01: Collect Q/R consequence and U/V sannsynlighet begrunnelser across each 10-row block.
# 2026-07-08: Forside values may appear on rows below header (not only same row).
# 2026-07-08: Forside scan bounded to rows 1–33 and columns C–P.
# 2026-07-08: Collect verdi(er) from all 10 block rows (D may be on any line).
# 2026-07-08: Large «risikovurderingsverktøy» Excel import – Risikovurdering 10-row blocks.

import unicodedata
from datetime import date, datetime

from django.db import transaction
from openpyxl.utils import column_index_from_string

from systemoversikt.models import RiskAction, RiskScenario
from systemoversikt.risk_criteria import get_active_criteria
from systemoversikt.risk_import import (
	ImportResult,
	_coerce_level,
	_is_placeholder_tiltak,
	_normalize_header,
	_title_from_filename,
)
from systemoversikt.risk_membership import create_risk_scope

DATA_START_ROW = 5
BLOCK_SIZE = 10
_HEADER_ROWS = (4, 5)
_HEADER_SKIP = frozenset({'riskid', 'scenariobeskrivelse'})
_KIT_WORD_TO_CODE = (
	('konfidensialitet', 'K'),
	('integritet', 'I'),
	('tilgjengelighet', 'T'),
)
_KIT_FIELD_MAX_LENGTH = 50

_COL = {
	'risk_id': column_index_from_string('B'),
	'scenario': column_index_from_string('C'),
	'verdi': column_index_from_string('D'),
	'kit': column_index_from_string('E'),
	'trussel': column_index_from_string('G'),
	'trusselnivaa': column_index_from_string('H'),
	'sarbarhet': column_index_from_string('I'),
	'eks_tiltak': column_index_from_string('O'),
	'kons_begrunnelse': column_index_from_string('Q'),
	'konsekvens': column_index_from_string('R'),
	'sanns_begrunnelse': column_index_from_string('S'),
	'sannsynlighet': column_index_from_string('T'),
	'tiltak': column_index_from_string('W'),
	'ansvarlig': column_index_from_string('AB'),
	'frist': column_index_from_string('AE'),
	'konsekvens_etter': column_index_from_string('AF'),
	'sannsynlighet_etter': column_index_from_string('AG'),
}

_FORSIDE_SKIP_PREFIXES = (
	'her kan det',
	'hvis du',
	'med risikoeier',
	'nb!',
	'(',
)

_FORSIDE_MAX_ROW = 33
_FORSIDE_VALUE_COL_START = column_index_from_string('C')
_FORSIDE_VALUE_COL_END = column_index_from_string('P')
_FORSIDE_VALUE_COLS = range(_FORSIDE_VALUE_COL_START, _FORSIDE_VALUE_COL_END + 1)


def _fold_header(text):
	normalized = unicodedata.normalize('NFKD', text)
	ascii_text = normalized.encode('ascii', 'ignore').decode('ascii')
	return ascii_text.lower().strip()


def _cell_val(ws, row, col_key, colmap=None):
	cols = colmap if colmap is not None else _COL
	return ws.cell(row, cols[col_key]).value


def _str_val(value):
	if value is None:
		return ''
	return str(value).strip()


def _kit_dimensjoner_from_excel(text):
	"""Fit stor-mal KIT prose into RiskScenario.kit_dimensjoner (K, I, T; max 50)."""
	raw = _str_val(text)
	if not raw:
		return ''
	folded = _fold_header(raw)
	codes = []
	for word, code in _KIT_WORD_TO_CODE:
		if word in folded and code not in codes:
			codes.append(code)
	if codes:
		return ', '.join(codes)
	if len(raw) <= _KIT_FIELD_MAX_LENGTH:
		return raw
	return raw[:_KIT_FIELD_MAX_LENGTH]


def _is_placeholder_tiltak_cell(text):
	"""Skip empty/xx-style stubs in eksisterende/planlagte tiltak columns only."""
	return _is_placeholder_tiltak(text)


def _parse_risk_num(value):
	"""Return 1-based risk id integer, or None for headers/empty/non-numeric."""
	if value is None or isinstance(value, bool):
		return None
	if isinstance(value, (int, float)):
		n = int(value)
		if n == value and n >= 1:
			return n
		return None
	text = _str_val(value)
	if not text or _fold_header(text) in _HEADER_SKIP:
		return None
	try:
		n = int(float(text.replace(',', '.')))
	except (TypeError, ValueError):
		return None
	if n >= 1:
		return n
	return None


def _resolve_columns(ws):
	"""Overlay header-based column indexes on the legacy letter map (old .xlsx fallback)."""
	colmap = dict(_COL)
	assigned = set()
	max_col = max(ws.max_column or 1, 40)
	for col in range(1, max_col + 1):
		parts = []
		for header_row in _HEADER_ROWS:
			part = _normalize_header(ws.cell(header_row, col).value)
			if part:
				parts.append(part)
		if not parts:
			continue
		folded = _fold_header(' '.join(parts))
		if 'dummy' in folded:
			continue
		key = None
		if 'risikoreduserende tiltak' in folded:
			key = 'tiltak'
		elif 'konsekvens etter' in folded:
			key = 'konsekvens_etter'
		elif 'sannsynlighet etter' in folded:
			key = 'sannsynlighet_etter'
		elif 'beskrivelse av konsekvens' in folded:
			key = 'kons_begrunnelse'
		elif 'beskrivelse av sannsynlighet' in folded:
			key = 'sanns_begrunnelse'
		elif folded.startswith('konsekvensniva'):
			key = 'konsekvens'
		elif 'sannsynlighetsniva' in folded:
			key = 'sannsynlighet'
		elif folded == 'ansvarlig':
			key = 'ansvarlig'
		elif folded.startswith('frist'):
			key = 'frist'
		elif 'eksisterende tiltak' in folded:
			key = 'eks_tiltak'
		if key and key not in assigned:
			colmap[key] = col
			assigned.add(key)

	# Newer mal: Q/R share the consequence-description header; R is dimension text, not the level.
	q_col = colmap.get('kons_begrunnelse')
	if q_col and colmap.get('konsekvens') != q_col + 1:
		colmap['kons_detalj'] = q_col + 1
	# Newer mal: T/U/V share sannsynlighet-description header; T is question labels, U/V are dimension+text.
	s_col = colmap.get('sanns_begrunnelse')
	if s_col:
		dim_col = s_col + 1
		detalj_col = s_col + 2
		level_col = colmap.get('sannsynlighet')
		other_cols = {col for key, col in colmap.items() if key != 'sanns_begrunnelse'}
		if (
			level_col not in (s_col, dim_col, detalj_col)
			and dim_col not in other_cols
			and detalj_col not in other_cols
		):
			colmap['sanns_begrunnelse'] = dim_col
			colmap['sanns_detalj'] = detalj_col
	return colmap


def _find_data_start_row(ws, colmap):
	"""First row with a numeric RiskID (row 6 on newer xlsm; row 5 on older xlsx)."""
	risk_col = colmap['risk_id']
	scenario_col = colmap['scenario']
	scan_end = min(ws.max_row, DATA_START_ROW + BLOCK_SIZE * 2)
	for row in range(DATA_START_ROW, scan_end + 1):
		if _parse_risk_num(ws.cell(row, risk_col).value) is None:
			continue
		scenario = _fold_header(_str_val(ws.cell(row, scenario_col).value))
		if scenario in _HEADER_SKIP:
			continue
		return row
	return DATA_START_ROW


def _collect_sarbarheter(ws, start_row, colmap):
	"""All non-empty I-column strings in the block, one per line."""
	lines = []
	for offset in range(BLOCK_SIZE):
		text = _str_val(_cell_val(ws, start_row + offset, 'sarbarhet', colmap))
		if text:
			lines.append(text)
	return lines


def _collect_verdier(ws, start_row, colmap):
	"""All non-empty D-column verdi names in the block (any of the 10 rows)."""
	names = []
	for offset in range(BLOCK_SIZE):
		name = _str_val(_cell_val(ws, start_row + offset, 'verdi', colmap))
		if name:
			names.append(name)
	return names


def _collect_paired_begrunnelse(ws, start_row, colmap, dim_key, detalj_key):
	"""Join dimension + optional detail text across the 10-row block (Q/R or U/V)."""
	lines = []
	has_detalj = detalj_key in colmap
	for offset in range(BLOCK_SIZE):
		dim = _str_val(_cell_val(ws, start_row + offset, dim_key, colmap))
		detalj = ''
		if has_detalj:
			detalj = _str_val(_cell_val(ws, start_row + offset, detalj_key, colmap))
		if dim and detalj:
			lines.append('%s: %s' % (dim, detalj))
		elif dim:
			lines.append(dim)
		elif detalj:
			lines.append(detalj)
	return '\n'.join(lines)


def _collect_konsekvens_begrunnelse(ws, start_row, colmap):
	"""Join Q (dimension) and optional R (text) across the 10-row block."""
	return _collect_paired_begrunnelse(ws, start_row, colmap, 'kons_begrunnelse', 'kons_detalj')


def _collect_sanns_begrunnelse(ws, start_row, colmap):
	"""Join U (dimension) and optional V (text); T question labels are skipped via column map."""
	return _collect_paired_begrunnelse(ws, start_row, colmap, 'sanns_begrunnelse', 'sanns_detalj')


def _coerce_frist(value):
	if isinstance(value, datetime):
		return value.date()
	if isinstance(value, date):
		return value
	return None


def _build_verdi_lookup(workbook):
	"""Verdivurdering col B name -> Beregnet verdi (1–5) from col J."""
	if 'Verdivurdering' not in workbook.sheetnames:
		return {}
	ws = workbook['Verdivurdering']
	lookup = {}
	for row in range(2, ws.max_row + 1):
		name = _str_val(ws.cell(row, 2).value)
		if not name:
			continue
		level = ws.cell(row, 10).value
		try:
			n = int(level)
			if 1 <= n <= 5:
				lookup[name] = n
		except (TypeError, ValueError):
			pass
	return lookup


def _build_trussel_lookup(workbook):
	"""Trusselvurdering col B name -> level text (manual P overrides auto O)."""
	if 'Trusselvurdering' not in workbook.sheetnames:
		return {}
	ws = workbook['Trusselvurdering']
	lookup = {}
	for row in range(3, ws.max_row + 1):
		name = _str_val(ws.cell(row, 2).value)
		if not name:
			continue
		manual = _str_val(ws.cell(row, 16).value)
		auto = _str_val(ws.cell(row, 15).value)
		level_text = manual or auto
		if level_text and level_text not in ('#N/A', '#REF!'):
			lookup[name] = level_text
	return lookup


def _konsekvens_label(level):
	if not level:
		return ''
	return get_active_criteria().konsekvens_lookup_label(level) or str(level)


def _parse_risikovurdering_data(workbook):
	"""Map risk block number (1-based) -> numeric level dict from hidden data sheet."""
	if 'Risikovurdering_data' not in workbook.sheetnames:
		return {}
	ws = workbook['Risikovurdering_data']
	result = {}
	for row in range(3, ws.max_row + 1, BLOCK_SIZE):
		risk_ref = ws.cell(row, 1).value
		if not risk_ref or str(risk_ref).strip() in ('0', ''):
			continue
		digits = ''.join(c for c in str(risk_ref) if c.isdigit())
		if not digits:
			continue
		block_num = int(digits)
		result[block_num] = {
			'sannsynlighet': ws.cell(row, 3).value,
			'konsekvens': ws.cell(row, 4).value,
			'sannsynlighet_etter': ws.cell(row, 6).value,
			'konsekvens_etter': ws.cell(row, 7).value,
		}
	return result


def _is_forside_boilerplate(label):
	lower = label.lower()
	return any(lower.startswith(p) for p in _FORSIDE_SKIP_PREFIXES)


def _collect_forside_row_values(ws, row):
	values = []
	for col in _FORSIDE_VALUE_COLS:
		v = _str_val(ws.cell(row, col).value)
		if v:
			values.append(v)
	return values


def _extract_forside_beskrivelse(workbook):
	"""Best-effort dump of Forside label/value pairs into scope beskrivelse."""
	if 'Forside' not in workbook.sheetnames:
		return ''
	ws = workbook['Forside']
	lines = []
	row = 1
	while row <= _FORSIDE_MAX_ROW:
		label = _str_val(ws.cell(row, 2).value)
		if not label or _is_forside_boilerplate(label):
			row += 1
			continue

		values = _collect_forside_row_values(ws, row)
		scan_row = row + 1
		while scan_row <= _FORSIDE_MAX_ROW:
			next_label = _str_val(ws.cell(scan_row, 2).value)
			if next_label and not _is_forside_boilerplate(next_label):
				break
			values.extend(_collect_forside_row_values(ws, scan_row))
			scan_row += 1

		if values:
			lines.append('%s: %s' % (label.rstrip(':'), ', '.join(values)))
		else:
			lines.append(label.rstrip(':'))
		row = scan_row
	return '\n'.join(lines).strip()


def _compose_uonsket_hendelse(scenario_text, verdier, trussel, trussel_level_text,
		verdi_lookup, trussel_lookup, warnings, risk_id):
	parts = []
	base = _str_val(scenario_text)
	if base:
		parts.append(base)

	verdi_bits = []
	for name in verdier:
		level = verdi_lookup.get(name)
		label = _konsekvens_label(level) if level else ''
		if label:
			verdi_bits.append('%s (%s)' % (name, label.lower()))
		else:
			verdi_bits.append(name)
			if name not in verdi_lookup:
				warnings.append('%s: ukjent verdi %r i Verdivurdering' % (risk_id, name))
	if verdi_bits:
		parts.append('Berører verdi(er): %s.' % ', '.join(verdi_bits))

	trussel_name = _str_val(trussel)
	if trussel_name:
		level_text = _str_val(trussel_level_text)
		if not level_text:
			level_text = trussel_lookup.get(trussel_name, '')
		if level_text:
			parts.append('Trussel: %s (%s).' % (trussel_name, level_text.lower()))
		else:
			parts.append('Trussel: %s.' % trussel_name)

	return ' '.join(parts).strip()


def _risk_id_from_block(risk_num):
	try:
		n = int(risk_num)
	except (TypeError, ValueError):
		n = risk_num
	return 'R%s' % n


def import_large_risk_workbook(workbook, user, source_filename):
	"""
	Import large «risikovurderingsverktøy» template (Risikovurdering sheet).
	Rolls back on error (atomic transaction).
	"""
	if 'Risikovurdering' not in workbook.sheetnames:
		raise ValueError('Mangler ark «Risikovurdering» i Excel-filen.')

	warnings = []
	ws = workbook['Risikovurdering']
	colmap = _resolve_columns(ws)
	data_start_row = _find_data_start_row(ws, colmap)
	verdi_lookup = _build_verdi_lookup(workbook)
	trussel_lookup = _build_trussel_lookup(workbook)
	data_by_block = _parse_risikovurdering_data(workbook)
	forside_text = _extract_forside_beskrivelse(workbook)
	if forside_text:
		warnings.append('Forside-tekst er limt inn i beskrivelse – vurder manuelt.')

	with transaction.atomic():
		scope = create_risk_scope(
			user,
			title=_title_from_filename(source_filename),
			beskrivelse=forside_text,
			sist_revidert=date.today(),
			source_filename=source_filename or '',
		)
		scenario_count = 0
		action_count = 0
		rekkefolge = 0

		for start_row in range(data_start_row, ws.max_row + 1, BLOCK_SIZE):
			risk_num = _parse_risk_num(_cell_val(ws, start_row, 'risk_id', colmap))
			if risk_num is None:
				continue
			scenario_text = _str_val(_cell_val(ws, start_row, 'scenario', colmap))
			if not scenario_text or _fold_header(scenario_text) in _HEADER_SKIP:
				continue
			rekkefolge += 1
			risk_id = _risk_id_from_block(risk_num)
			data_row = data_by_block.get(risk_num, {})

			konsekvens = _coerce_level(
				_cell_val(ws, start_row, 'konsekvens', colmap),
				data_row.get('konsekvens'),
				'konsekvens',
			)
			sannsynlighet = _coerce_level(
				_cell_val(ws, start_row, 'sannsynlighet', colmap),
				data_row.get('sannsynlighet'),
				'sannsynlighet',
			)
			konsekvens_etter = _coerce_level(
				_cell_val(ws, start_row, 'konsekvens_etter', colmap),
				data_row.get('konsekvens_etter'),
				'konsekvens',
			)
			sannsynlighet_etter = _coerce_level(
				_cell_val(ws, start_row, 'sannsynlighet_etter', colmap),
				data_row.get('sannsynlighet_etter'),
				'sannsynlighet',
			)

			if konsekvens is None and _str_val(_cell_val(ws, start_row, 'konsekvens', colmap)):
				warnings.append('%s: ukjent konsekvens %r' % (
					risk_id, _cell_val(ws, start_row, 'konsekvens', colmap)))
			if sannsynlighet is None and _str_val(_cell_val(ws, start_row, 'sannsynlighet', colmap)):
				warnings.append('%s: ukjent sannsynlighet %r' % (
					risk_id, _cell_val(ws, start_row, 'sannsynlighet', colmap)))

			verdier = _collect_verdier(ws, start_row, colmap)

			uonsket_hendelse = _compose_uonsket_hendelse(
				scenario_text,
				verdier,
				_cell_val(ws, start_row, 'trussel', colmap),
				_cell_val(ws, start_row, 'trusselnivaa', colmap),
				verdi_lookup,
				trussel_lookup,
				warnings,
				risk_id,
			)

			sarbarheter = _collect_sarbarheter(ws, start_row, colmap)
			scenario = RiskScenario.objects.create(
				scope=scope,
				risk_id=risk_id,
				uonsket_hendelse=uonsket_hendelse,
				kit_dimensjoner=_kit_dimensjoner_from_excel(
					_cell_val(ws, start_row, 'kit', colmap)),
				arsaker_svakheter='\n'.join(sarbarheter),
				konsekvens_nivaa=konsekvens,
				sannsynlighet_nivaa=sannsynlighet,
				konsekvens_begrunnelse=_collect_konsekvens_begrunnelse(ws, start_row, colmap),
				sannsynlighetsbegrunnelse=_collect_sanns_begrunnelse(ws, start_row, colmap),
				risikobehandling='',
				konsekvens_etter=konsekvens_etter,
				sannsynlighet_etter=sannsynlighet_etter,
				rekkefolge=rekkefolge,
			)
			scenario_count += 1

			for offset in range(BLOCK_SIZE):
				row = start_row + offset
				eks = _cell_val(ws, row, 'eks_tiltak', colmap)
				if not _is_placeholder_tiltak_cell(eks):
					text = _str_val(eks)
					if text:
						action = RiskAction.objects.create(
							scope=scope,
							beskrivelse=text,
							kilde='parsed',
							status='utfort',
						)
						action.scenarios.add(scenario)
						action_count += 1

				plan = _cell_val(ws, row, 'tiltak', colmap)
				if not _is_placeholder_tiltak_cell(plan):
					text = _str_val(plan)
					if text:
						action = RiskAction.objects.create(
							scope=scope,
							beskrivelse=text,
							ansvarlig=_str_val(_cell_val(ws, row, 'ansvarlig', colmap)),
							frist=_coerce_frist(_cell_val(ws, row, 'frist', colmap)),
							kilde='parsed',
							status='besluttet',
						)
						action.scenarios.add(scenario)
						action_count += 1

	return ImportResult(
		scope=scope,
		scenario_count=scenario_count,
		action_count=action_count,
		warnings=warnings,
		format='large',
	)
